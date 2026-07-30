from flask import Flask, send_file, jsonify
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill
import re
import time
from io import BytesIO
import os

app = Flask(__name__)

CHANNELS = {
    "Александр Соколовский": "UCaR6XjSJJsLbKN3n6VYsGKw",
    "ВПИСКА": "UCj7bSQWlq2O4lhGxGll5SUA"
}

BRANDS_RU = [
    "Яндекс Путешествия",
    "Суточно.ру",
    "Островок",
    "Твил",
    "Авито Путешествия",
    "Ozon Travel",
    "Отелло",
    "Альфа Путешествия",
    "Т-Путешествия",
    "Trip.com",
    "Fun&Sun",
    "Aviasales",
    "Туту.ру",
    "OneTwoTrip"
]

BRANDS_EN = [
    "Yandex Travel",
    "Sutonochno",
    "Ostrovok",
    "Twil",
    "Avito Travel",
    "Ozon Travel",
    "Otello",
    "Alfa Travel",
    "T-Travel",
    "Trip.com",
    "Fun&Sun",
    "Aviasales",
    "Tutu.ru",
    "OneTwoTrip"
]

ALL_BRANDS = BRANDS_RU + BRANDS_EN

def get_channel_videos(channel_id):
    url = f"https://www.youtube.com/c/{channel_id}/videos"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    }
    try:
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()
        return response.text
    except Exception as e:
        print(f"Ошибка: {e}")
        return None

def extract_video_info(html_content):
    videos = []
    video_links = re.findall(r'/watch\?v=([a-zA-Z0-9_-]{11})', html_content)
    for video_id in video_links:
        videos.append({
            "id": video_id,
            "url": f"https://www.youtube.com/watch?v={video_id}"
        })
    return videos

def get_video_details(video_url):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    }
    try:
        response = requests.get(video_url, headers=headers, timeout=10)
        response.raise_for_status()
        soup = BeautifulSoup(response.content, 'html.parser')
        title = soup.find('meta', {'name': 'title'})
        title_text = title['content'] if title else "Unknown"
        description = soup.find('meta', {'name': 'description'})
        description_text = description['content'] if description else ""
        return title_text, description_text
    except Exception as e:
        return None, None

def find_brand_mentions(title, description, brands):
    text = f"{title} {description}".lower()
    found_brands = []
    for brand in brands:
        if brand.lower() in text:
            found_brands.append(brand)
    return found_brands

def save_to_excel(results):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Результаты"
    headers = ["Блогер", "Бренд", "Ссылка на видео", "Название видео"]
    ws.append(headers)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for result in results:
        ws.append([result["channel"], result["brand"], result["url"], result["title"]])
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 40
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

@app.route('/')
def index():
    return jsonify({"status": "API запущена", "message": "Перейди на /analyze для анализа"})

@app.route('/analyze')
def analyze():
    try:
        all_results = []
        for channel_name, channel_id in CHANNELS.items():
            html = get_channel_videos(channel_id)
            if not html:
                continue
            videos = extract_video_info(html)
            for idx, video in enumerate(videos[:50]):
                time.sleep(0.5)
                title, description = get_video_details(video['url'])
                if title and description:
                    brands_found = find_brand_mentions(title, description, ALL_BRANDS)
                    for brand in brands_found:
                        all_results.append({
                            "channel": channel_name,
                            "brand": brand,
                            "url": video['url'],
                            "title": title
                        })
        if all_results:
            excel_file = save_to_excel(all_results)
            return send_file(
                excel_file,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='polymarket_brands.xlsx'
            )
        else:
            return jsonify({"status": "error", "message": "Упоминаний не найдено"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
